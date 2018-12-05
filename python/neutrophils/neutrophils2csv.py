import argparse
import json
import sys, os
from collections import defaultdict, Counter

from lxml import etree
from xml.dom import minidom

sys.path.insert(0, str(os.path.dirname(os.path.realpath(__file__))) + "/../")


from neutrophils.allowedSentences import AllowedSentences



"""

no eviction of stupid sentences

relEvCount 8827
evCount 20634
evPMID 5388
evSents 7586
PMIDs 4390
PMCs 1219
"""


def findContextInfo(inContext, allowedSentIDs, acceptDocWide=False):

    foundContexts = set()
    contextDistances = {}
    contextWords = {}

    if acceptDocWide:
        return foundContexts, contextWords, contextDistances

    for termContext in inContext:

        if termContext == None:
            continue

        for x in termContext['evidences']:
            if x[0] == sentid:
                contextElem = (termContext['termid'], termContext['termname'], x[0], x[1], x[2])
                foundContexts.add(contextElem)

                if contextElem in contextDistances:
                    if allowedSentIDs[x[0]] < contextDistances[contextElem]:
                        contextDistances[contextElem] = allowedSentIDs[x[0]]

                        if x[0] == sentid:
                            contextWords[contextElem] = makeSubstr(sentence, x[1], x[2])
                else:
                    contextDistances[contextElem] = allowedSentIDs[x[0]]

                    if x[0] == sentid:
                        contextWords[contextElem] = makeSubstr(sentence, x[1], x[2])

    if len(foundContexts) == 0:
        for termContext in inContext:

            if termContext == None:
                continue

            for x in termContext['evidences']:
                if x[0] in allowedSentIDs:

                    contextElem = (termContext['termid'], termContext['termname'], x[0], x[1], x[2])
                    foundContexts.add(contextElem)

                    if contextElem in contextDistances:

                        if allowedSentIDs[x[0]] < contextDistances[contextElem]:
                            contextDistances[contextElem] = allowedSentIDs[x[0]]

                            if x[0] == sentid:
                                contextWords[contextElem] = makeSubstr(sentence, x[1], x[2])


                    else:
                        contextDistances[contextElem] = allowedSentIDs[x[0]]

                        if x[0] == sentid:
                            contextWords[contextElem] = makeSubstr(sentence, x[1], x[2])

    return foundContexts, contextWords, contextDistances



import requests

from textmining.SentenceID import SentenceID





def extractPMCInfo(handle):
    root = etree.parse(handle)

    lpmid2info = {}
    reviewIDs = set()
    for article in root.findall('.//article/'):

        articleIDs = [x.text for x in article.findall(".//article-id[@pub-id-type='pmc']")]

        if len(articleIDs) == 0:
            continue

        link = "https://www.ncbi.nlm.nih.gov/pmc/articles/PMC" + articleIDs[0] if len(articleIDs) > 0 else ""

        articleTitles = [x for x in article.findall('.//title-group')]
        articleTitle = ""

        for titleGroup in articleTitles:
            tgText = "".join([x.strip() for x in titleGroup.itertext()])
            articleTitle += tgText

        articleJournals = article.findall('.//journal-title')
        articleJournal = "" if len(articleJournals) == 0 else articleJournals[0].text

        articleAuthors = [x for x in article.findall(".//contrib[@contrib-type='author']")]

        authors = [",".join([x for x in au.itertext() if len(x.strip()) > 0]) for au in articleAuthors]
        authorString = ""

        if len(authors) > 0:
            authorString = "; ".join(authors)

        articleYears = article.findall('.//pub-date/year')
        articleYear = -1 if len(articleYears) == 0 else articleYears[0].text

        lpmid2info["PMC" + articleIDs[0]] = (link, articleTitle, articleJournal, articleYear, authorString)

        if "REVIEW" in articleTitle.upper():
            reviewIDs.add("PMC"+articleIDs[0])

    return lpmid2info, reviewIDs

def extractPMIDInfo(handle):

    record = Entrez.read(handle)

    lpmid2info = {}
    reviewIDs = set()

    for article in record['PubmedArticle']:

        pubmedID = str(article['PubmedData']['ArticleIdList'][0]) if len(
            article['PubmedData']['ArticleIdList']) > 0 else "-1"
        pubID = int(pubmedID)

        artInfo = article['MedlineCitation']['Article']
        articleTitle = artInfo['ArticleTitle']
        articleJournal = artInfo['Journal']['Title'] if 'Journal' in artInfo else ''

        artDate = artInfo["ArticleDate"]

        articleYear = -1

        if len(artDate) > 0:
            articleYear = artDate[0]["Year"]

        if articleYear == -1 and 'Journal' in artInfo:

            articleJournalInfo = artInfo["Journal"]
            if "JournalIssue" in articleJournalInfo:
                if 'PubDate' in articleJournalInfo['JournalIssue']:
                    if 'Year' in articleJournalInfo['JournalIssue']['PubDate']:
                        articleYear = articleJournalInfo['JournalIssue']['PubDate']['Year']


        articleAuthors = ""
        if "AuthorList" in artInfo and len(artInfo["AuthorList"]) > 0:

            allauthors = []
            for author in artInfo ["AuthorList"]:
                tlauthor = author["LastName"] if "LastName" in author else ""
                tfauthor = author["ForeName"] if "ForeName" in author else ""

                if tlauthor == "" and tfauthor == "":
                    print("Stupid authors", pubmedID)
                else:
                    allauthors.append(tlauthor + ", " + tfauthor)

            articleAuthors = "; ".join(allauthors)


        if "PublicationTypeList" in artInfo and "Review" in artInfo['PublicationTypeList']:
            reviewIDs.add(pubmedID)


        infotuple = ("https://www.ncbi.nlm.nih.gov/pubmed/" + pubmedID, articleTitle, articleJournal, articleYear, articleAuthors)

        lpmid2info[pubmedID] = infotuple

    return lpmid2info, reviewIDs



def getResultHandle(dbName, allIDs):

    epostResult = Entrez.read(Entrez.epost(dbName, id=",".join(allIDs)))
    webEnv = epostResult['WebEnv']
    queryKey = epostResult['QueryKey']

    handle = Entrez.efetch(db=dbName, webenv=webEnv, query_key=queryKey, retmode='XML')

    return handle


#thandle = getResultHandle("pubmed", ['27175518'])
#print(extractPMIDInfo(thandle))

#exit(0)
#
#thandle = getResultHandle("pmc", ['4902075'])
#print(extractPMCInfo(thandle))

parser = argparse.ArgumentParser(description='db query', add_help=False)
parser.add_argument('-o', '--output', type=argparse.FileType("w"), help='outfile', default=None, required=False)
parser.add_argument('-t', '--output-text', type=argparse.FileType("w"), help='outfile', default=sys.stdout, required=False)
parser.add_argument('--organ', type=str, help='outfile', default=None, required=False)

args = parser.parse_args()

fetchAddData = False
fetchSentences=False



from Bio import Entrez

Entrez.email = "joppich@bio.ifi.lmu.de"

maxSentDist = 3
query = {
        "elements": [{'group': 'NEUTROPHIL', 'name': 'PMN', 'termid': 'PMN'}, {'group': 'NEUTROPHIL', 'name': 'neutrophils', 'termid': 'neutrophils'}],
        "sentences": str(fetchSentences),
        "obolevel": 1,
        "messenger_obolevel": 1,
        "sentence_distance": maxSentDist
    }

organQuery = []
if args.organ != None:
    organQuery = [{"group": "organs", "name": args.organ, "termid": args.organ}]

    query['organs'] = organQuery
    print("Set organs", organQuery, file=sys.stderr)

for x in query:
    print(x, query[x])

r = requests.post("http://localhost:65522/query", data=json.dumps(query))

res = json.loads(r.content.decode())

sep = "\t"

allPMIDs = set()
allPMC = set()

print("received", len(res['rels']), "relations", file=sys.stderr)

for rel in res['rels']:

    for ev in rel['evidences']:

        docid = ev['docid']

        if docid == None:
            continue

        if docid.startswith("PMC"):
            allPMC.add(docid.replace('PMC', ""))
        else:
            allPMIDs.add(docid)





pmid2info = {}

reviewArticles =set()

if fetchAddData:

    for dbName, allIDs in [('pmc', allPMC), ('pubmed', allPMIDs)]:

        handle = getResultHandle(dbName, allIDs)

        allAuthorArticles = []

        if dbName == 'pmc':

            lpmid2info, lreview = extractPMCInfo(handle)
            reviewArticles = reviewArticles.union(lreview)

            for x in lpmid2info:
                pmid2info[x] = lpmid2info[x]




        else:
            lpmid2info, lreview = extractPMIDInfo(handle)

            reviewArticles = reviewArticles.union(lreview)

            for x in lpmid2info:
                pmid2info[x] = lpmid2info[x]


elemcount = 0

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

header = ["PMID", "Sent ID", "Message ID", "Message Ontology ID", "Effect ID", "Effect Ontology ID", "Sentence", "Verb Structure", "Left ID", "Left Ontology ID",
          "Right ID", "Right Ontology ID", "Link", "Title", "Journal", "Year", "Authors", "Verb", "Stack", "Relex", "Conj", "EFF Dist", "MESS Dist", "LWORD", "RWORD", "EWORD", "MWORD"]
print(header, sep=sep, file=args.output_text)
ws.append(header)

relEvCount = 0
evCount = 0
evPMID = set()
evPMC=set()
evSent = set()
allowedSentsGen = AllowedSentences(maxSentDist)

dist2tuples = defaultdict(lambda: Counter())
minDistForDoc = {}

def makeSubstr(sent, s, e):

    if len(sent) <= e:
        return ""

    return sent[s:e]


for rel in res['rels']:

    for ev in rel['evidences']:

        docid = ev['docid']

        if docid in reviewArticles:
            print("Skipping review", docid)

        sentid = ev['rel_sentence']
        sentence = ev.get("sentence", "")

        aSent = sentid.split(".")
        aSentNum = int(aSent[-1])

        aSent = aSent[0:2]

        # remove citations/references
        if aSent[0].startswith("PMC") and aSent[1] == '4':
            continue

        if len(sentence) > 0:
            if "///" in sentence:
                continue



        allowedSentIDs = allowedSentsGen.getDistanceDictBySentID(sentid)

        verbdir = ev['rel_direction_verb']

        lid = ev['lid']
        loid = ev['lontid']

        rid = ev['rid']
        roid = ev['rontid']

        trusts = (ev['trust']['verb'], ev['trust']['stack'], ev['trust']['relex'], ev['trust']['conj'])

        #if len([x for x in trusts[0:3] if x > 0]) == 0:
        #    continue

        textLeft = ""
        textRight = ""

        if len(sentence) > 0:
            textLeft = sentence[ev['lpos'][0]:ev['lpos'][1]]
            textRight = sentence[ev['rpos'][0]:ev['rpos'][1]]

        effect = res['pmidinfo']['categories'].get(docid, [None])
        message = res['pmidinfo']['messengers'].get(docid, [None])
        organ = res['pmidinfo']['organs'].get(docid, [None])


        foundOrgans = set()
        organWords = {}
        organDistances = {}
        if len(organQuery) > 0: # required specific organ

            if len(organ) == 0:

                print("Organ Ejected", docid)
                continue


        foundEffects = set()
        effectDistances = {}
        effectWords = {}

        if effect != None:
            foundEffects, effectWords, effectDistances = findContextInfo(effect, allowedSentIDs)

        foundMessages = set()
        messageWords = {}
        messageDistances = {}

        if message != None:
            foundMessages, messageWords, messageDistances = findContextInfo(message, allowedSentIDs)


        trustStr = sep.join([str(x) for x in trusts])

        infoTuple = pmid2info.get(docid, ("", "","",-2, ""))

        infoTuple = [str(x) for x in infoTuple]

        elemcount += len(foundEffects) * len(foundMessages)

        #if len(foundEffects) == 0 or len(foundMessages) == 0:
        #    print("INC EFF MESS", len(foundEffects), len(foundMessages))

        relEvCount += 1


        for messageElem in foundMessages:
            for effectElem in foundEffects:
                mDist = messageDistances[messageElem]
                eDist = effectDistances[effectElem]

                messageWord = messageWords.get(messageElem, "")
                effectWord = effectWords.get(effectElem, "")


                if messageElem[2] == effectElem[2]:
                    messageInterval = (messageElem[3], messageElem[4])
                    effectInterval = (effectElem[3], effectElem[4])

                    overlap = max([messageInterval[0], effectInterval[0]]) <= min([messageInterval[1], effectInterval[1]])

                    if overlap:
                        #print("overlap")
                        #print(messageInterval, messageWord)
                        #print(effectInterval, effectWord)
                        continue


                docDistance = max([abs(eDist), abs(mDist)])
                maxDocDistance = min([minDistForDoc.get(docid, 10), docDistance])

                minDistForDoc[docid] = maxDocDistance

                entryTuple = (lid, rid, messageElem[1], effectElem[1])

                dist2tuples[docDistance][entryTuple] += 1

                allData = [docid, sentid] + [messageElem[0], messageElem[1], effectElem[0], effectElem[1]] + [sentence, verbdir, lid, loid, rid, roid]+ list(infoTuple) + list(trusts)\
                          + [str(eDist), str(mDist)] + [textLeft, textRight, effectWord, messageWord]
                ws.append(allData   )

                ws.cell(row=ws._current_row, column=13).hyperlink = infoTuple[0]
                ws.cell(row=ws._current_row, column=13).value = infoTuple[0]
                ws.cell(row=ws._current_row, column=13).style = "Hyperlink"

                print(allData, sep=sep, file=args.output_text)

                evCount += 1
                if docid.startswith("PMC"):
                    evPMC.add(docid)
                else:
                    evPMID.add(docid)

                evSent.add(sentid)


ndist2tuples = {}

for x in dist2tuples:
    ndist2tuples[x] = dist2tuples[x]

print("ndist2tuples =",ndist2tuples)
print("minDistForDoc = ", minDistForDoc)


with open("/mnt/c/Users/mjopp/Desktop/minimal_dist_per_doc.tsv", "w") as fout:
    for docid in minDistForDoc:
        infoTuple = pmid2info.get(docid, ("", "","",-2, ""))

        fout.write("\t".join([str(docid), str(minDistForDoc[docid])] + list([str(x) for x in infoTuple])) + "\n")


print("relEvCount", relEvCount)
print("evCount", evCount)
print("evPMID", len(evPMID))
print("evPMID", len(evPMC))
print("evSents", len(evSent))
print("PMIDs", len(allPMIDs))
print("PMCs", len(allPMC))
# Save the file

if args.output != None:
    wb.save(args.output.name)

#print(elemcount, file=sys.stderr)